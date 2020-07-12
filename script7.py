import openpyxl as op
from datetime import datetime
import pandas as pd
import os
path=os.getcwd()
wb = op.load_workbook(r"template.xlsx".format(path))   #for knn file
temp_sheet = wb["Sheet1"]
sheet2 = wb.create_sheet("Below -4")

def findingrange(gelensheet):
    hassan=0
    for index, row in enumerate(gelensheet.iter_rows()):
        hassan+=1
        #print(hassan)
    return hassan
def pastedata(a,gelen):
    print("module: paste()")
    #print(gelen)
    limitpaste=len(gelen)
    for i in range(0,limitpaste):
        yapistir=gelen[i]
        paste=temp_sheet.cell(row=a,column=i+1)
        paste.value = yapistir
def pastedata2(a,gelen):
    print("module: paste()")
    #print(gelen)
    limitpaste=len(gelen)
    for i in range(0,limitpaste):
        yapistir=gelen[i]
        paste=sheet2.cell(row=a,column=i+1)
        paste.value = yapistir
filenames=[]
def makefiledays():
    ##############  MAIN FILE  ###################
    colnames=['Timestamp','Open','High','Low','Close']
    csvread= pd.read_csv("the super boring stuf1.csv", names=colnames,index_col=False, skiprows=1)
    
    times=csvread.Timestamp.tolist()
    df=csvread.to_numpy()
    #print(times)
    ##############  PASTE CSV FILE  ###################
    def pastedatacsv(gelenlist):
        print(gelenlist)
        tostr=gelenlist
        limit=len(tostr)
        string=''
        for i in range(0,limit):
            if i==limit-1:
                string=string+str(tostr[i])
            else:
                string=string+str(tostr[i])+','    
        print('pstd',string)
        File_object.write(string+'\n')
    
    days=[]
    daysdiff=[]
    
    for i in df:
        date=datetime.strptime(i[0], "%Y:%m:%d-%H:%M:%S")
        days.append(date.day)
        for i in days:
            if i not in daysdiff:
                daysdiff.append(i)
        #print(daysdiff)
    for k in daysdiff:
        oneday=[]
        for get in df:
            timestamp=datetime.strptime(get[0], "%Y:%m:%d-%H:%M:%S")
            #print(get)
            if((timestamp.day)==k):
                #print(timestamp)
                oneday.append(get)     
        File_object = open(r"{}\files\day{}.csv".format(path,k),"w+")
        for i in oneday:
            pastedatacsv(i)            
        File_object.close()
        filenames.append('day{}'.format(k))
        print('done making the file, now continuing')
        print('----------------------')
    for i in filenames:
        colnames=['Timestamp','Open','High','Low','Close']
        csvread= pd.read_csv(r"{}\files\{}.csv".format(path,i), names=colnames,index_col=False)
        #times=csvread,Timestamp.tolist()
        df=csvread.to_numpy()
        #print(df)
        # timesinfile=[]
        # for entry in df:
        #     timestamp=datetime.strptime(entry[0], "%Y:%m:%d-%H:%M:%S")
        #     timesinfile.append(timestamp)
        # print(timesinfile)
        a=1
        yaz=['Input','','','','','Results']
        pastedata(a,yaz)
        a+=1
        yaz2=['Timestamp','Open','High','Low','Close']
        pastedata(a,yaz2)
        a=1
        yaz=['Input','','','','','Results']
        pastedata2(a,yaz)
        a+=1
        yaz2=['Timestamp','Open','High','Low','Close']
        pastedata2(a,yaz2)
        for entry in df:
            a+=1
            pastedata(a,entry)
            pastedata2(a,entry)
        wb.save(r"{}\files\{}.xlsx".format(path,i))  
makefiledays()

############## FOR LOW ]##################
# for i in filenames:
#     print(i,'Calculating differences, sorting the results and pasting in excel file (LOW)')
#     work=op.load_workbook(r"{}\files\{}.xlsx".format(path,i))
#     sheet=work['Sheet1']
#     sheetiki=work['Below 4']
#     b=0
#     for z in range (3,findingrange(sheet)+1):
#         readcelldate=sheet.cell(row=z,column=1).value
#         readcell=sheet.cell(row=z,column=4).value
#         c4=sheet.cell(row=2,column=7+b)
#         c4.value='{}Pairwise Diff'.format(readcelldate)
#         c4=sheetiki.cell(row=2,column=7+b)
#         c4.value='{}Pairwise Diff'.format(readcelldate)
#         for k in range (z+1,findingrange(sheet)+1):
#             try:   
#                 readag=sheet.cell(row=k,column=4).value
#                 diff=readcell-readag
#                 #print(diff)
#                 if(diff<4):
#                     #c+=1
#                     c4=sheetiki.cell(row=k,column=7+b)
#                     c4.value=diff
#                 else:
#                     c4=sheet.cell(row=k,column=7+b)
#                     c4.value=diff
#             except:
#                 pass
#         b+=1
#     work.save("{}-low.xlsx".format(i))
# ############## FOR HIGH##################

# for i in filenames:
#     print(i,'Calculating differences, sorting the results and pasting in excel file (HIGH)')
#     work=op.load_workbook(r"{}\files\{}.xlsx".format(path,i))
#     sheet=work['Sheet1']
#     sheetiki=work['Below 4']
#     sheetiki.title = 'Above -4'
#     b=0
#     for z in range (3,findingrange(sheet)+1):
#         readcelldate=sheet.cell(row=z,column=1).value
#         readcell=sheet.cell(row=z,column=3).value
#         c4=sheet.cell(row=2,column=7+b)
#         c4.value='{}Pairwise Diff'.format(readcelldate)
#         c4=sheetiki.cell(row=2,column=7+b)
#         c4.value='{}Pairwise Diff'.format(readcelldate)
#         for k in range (z+1,findingrange(sheet)+1):
#             try:   
#                 readag=sheet.cell(row=k,column=3).value
#                 diff=readcell-readag
#                 #print(diff)
#                 if(diff>(-4)):
#                     #c+=1
#                     c4=sheetiki.cell(row=k,column=7+b)
#                     c4.value=diff
#                 else:
#                     c4=sheet.cell(row=k,column=7+b)
#                     c4.value=diff
#             except:
#                 pass
#         b+=1
#     work.save("{}-high.xlsx".format(i))

from multiprocessing import Pool
def low(gelen):
    print(gelen,'Calculating differences, sorting the results and pasting in excel file (LOW)')
    work=op.load_workbook(r"{}\files\{}.xlsx".format(path,gelen))
    sheet=work['Sheet1']
    sheetiki=work['Below -4']
    b=0
    for z in range (3,findingrange(sheet)+1):
        readcelldate=sheet.cell(row=z,column=1).value
        readcell=sheet.cell(row=z,column=4).value
        c4=sheet.cell(row=2,column=7+b)
        c4.value='{}Pairwise Diff'.format(readcelldate)
        c4=sheetiki.cell(row=2,column=7+b)
        c4.value='{}Pairwise Diff'.format(readcelldate)
        for k in range (z+1,findingrange(sheet)+1):
            try:   
                readag=sheet.cell(row=k,column=4).value
                diff=readcell-readag
                #print(diff)
                if(diff<-3):
                    #c+=1
                    c4=sheetiki.cell(row=k,column=7+b)
                    c4.value=diff
                else:
                    c4=sheet.cell(row=k,column=7+b)
                    c4.value=diff
            except:
                pass
        b+=1
    work.save("{}-low.xlsx".format(gelen))
def high(gelen):    
    print(gelen,'Calculating differences, sorting the results and pasting in excel file (HIGH)')
    work=op.load_workbook(r"{}\files\{}.xlsx".format(path,gelen))
    sheet=work['Sheet1']
    sheetiki=work['Below -4']
    sheetiki.title = 'Above 4'
    b=0
    for z in range (3,findingrange(sheet)+1):
        readcelldate=sheet.cell(row=z,column=1).value
        readcell=sheet.cell(row=z,column=3).value
        c4=sheet.cell(row=2,column=7+b)
        c4.value='{}Pairwise Diff'.format(readcelldate)
        c4=sheetiki.cell(row=2,column=7+b)
        c4.value='{}Pairwise Diff'.format(readcelldate)
        for k in range (z+1,findingrange(sheet)+1):
            try:   
                readag=sheet.cell(row=k,column=3).value
                diff=readcell-readag
                #print(diff)
                if(diff>(4)):
                    #c+=1
                    c4=sheetiki.cell(row=k,column=7+b)
                    c4.value=diff
                else:
                    c4=sheet.cell(row=k,column=7+b)
                    c4.value=diff
            except:
                pass
        b+=1
    work.save("{}-high.xlsx".format(gelen))
import multiprocessing
if __name__ == '__main__':
    jobs = []
    years=['day24','day25','day26','day28','day29','day30','day1','day2','day3']
    #years=['day24','day25','day26','day28','day29','day30','day1']
    for i in years:
        p = multiprocessing.Process(target=low, args=(i,))
        jobs.append(p)
        p.start()
    # for k in years:
    #     p = multiprocessing.Process(target=high, args=(k,))
    #     jobs.append(p)
    #     p.start()





